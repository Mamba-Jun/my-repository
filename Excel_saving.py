import openpyxl
from openpyxl import Workbook, load_workbook
import os
import threading
from datetime import datetime
import time
import random

class Excel_operate:
    def __init__(self, filename='C:/Users/39174/Desktop/Excel20260202.xlsx', Sheetname='sheet'):
        self.filename = filename
        self.Sheetname = Sheetname
        self.workingsheet = f'{Sheetname}10'
        self.currentrow = 2
        self.sheet_counter = 10

    def initial(self):
        if os.path.exists(self.filename):
            self.workbook = load_workbook(self.filename)
        else:
            self.workbook = Workbook()

        if self.workingsheet not in self.workbook.sheetnames:
            self.worksheet = self.workbook.create_sheet(title=self.workingsheet)
        else:
            self.worksheet = self.workbook[self.workingsheet]
        
        if self.worksheet.max_row == 1:
            Headers=['time','Temp','humidity']
            for col, Header in enumerate(Headers, start = 1):
                self.worksheet.cell(row=1,column=col,value=Header)
                self.currentrow = 2
            self.workbook.save(self.filename)
            print(f'完毕')
        else: 
            self.currentrow = self.worksheet.max_row +1
        self.workbook.save(self.filename)

    def data_generation(self):
        current_time = datetime.now()
        Temperature = round(random.uniform(18.0, 30.0), 1)
        Humidity = round(random.uniform(30.0, 95.0), 1)
        return [current_time, Temperature, Humidity]
    
    def save_data(self):
        data = self.data_generation()
        if self.currentrow >= 30:
            self.new_sheet()
        else:
            for col, Value in enumerate(data, start=1):
                self.worksheet.cell(row=self.currentrow,column=col,value=Value)
            self.currentrow = self.currentrow + 1
        self.workbook.save(self.filename)

    def new_sheet(self):
        self.sheet_counter = self.sheet_counter + 1
        self.Sheetname = f'sheet{self.sheet_counter}' 
        # self.workbook.create_sheet(title=self.Sheetname)
        self.worksheet = self.workbook.create_sheet(title=self.Sheetname)
        Headers=['time','Temp','humidity']
        for col, Header in enumerate(Headers, start = 1):
            self.worksheet.cell(row=1,column=col,value=Header)
        self.currentrow = 2
        self.workbook.save(self.filename)


if __name__ == "__main__":
    excel = Excel_operate()
    excel.initial()
    i = 0
for i in range(0,40):
    time.sleep(1)
    print(f'已经完成{i}次')
    excel.save_data()